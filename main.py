import sys
import sqlite3
from pathlib import Path
from datetime import datetime

from configuracoes_impressora_ui import ConfiguracoesImpressoraPage

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QMessageBox,
    QFrame,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QDialog,
    QFormLayout,
    QDoubleSpinBox,
    QDialogButtonBox,
    QComboBox,
    QTabWidget,
    QFileDialog,
)

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, Alignment

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure

from database import Database


BASE_DIR = Path(__file__).resolve().parent
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(exist_ok=True)


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def preencher_tabela(tabela, dados):
    tabela.setRowCount(len(dados))

    for linha, registro in enumerate(dados):
        for coluna, valor in enumerate(registro):
            tabela.setItem(
                linha,
                coluna,
                QTableWidgetItem(
                    "" if valor is None else str(valor)
                )
            )


# ============================================================
# PRODUTO
# ============================================================

class ProdutoDialog(QDialog):

    def __init__(self, parent=None, produto=None):
        super().__init__(parent)

        self.setWindowTitle(
            "Cadastrar Produto"
            if produto is None
            else "Editar Produto"
        )

        layout = QFormLayout(self)

        self.codigo = QLineEdit()
        self.nome = QLineEdit()
        self.categoria = QLineEdit()

        self.preco_compra = QDoubleSpinBox()
        self.preco_compra.setMaximum(999999999)
        self.preco_compra.setDecimals(2)
        self.preco_compra.setPrefix("R$ ")

        self.preco_venda = QDoubleSpinBox()
        self.preco_venda.setMaximum(999999999)
        self.preco_venda.setDecimals(2)
        self.preco_venda.setPrefix("R$ ")

        self.estoque = QDoubleSpinBox()
        self.estoque.setMaximum(999999999)
        self.estoque.setDecimals(2)

        self.estoque_minimo = QDoubleSpinBox()
        self.estoque_minimo.setMaximum(999999999)
        self.estoque_minimo.setDecimals(2)

        layout.addRow("Código:", self.codigo)
        layout.addRow("Nome:", self.nome)
        layout.addRow("Categoria:", self.categoria)
        layout.addRow("Preço de compra:", self.preco_compra)
        layout.addRow("Preço de venda:", self.preco_venda)
        layout.addRow("Estoque:", self.estoque)
        layout.addRow("Estoque mínimo:", self.estoque_minimo)

        botoes = QDialogButtonBox(
            QDialogButtonBox.Save |
            QDialogButtonBox.Cancel
        )

        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)

        layout.addRow(botoes)

        if produto:
            self.codigo.setText(produto[1] or "")
            self.nome.setText(produto[2] or "")
            self.categoria.setText(produto[3] or "")
            self.preco_compra.setValue(float(produto[4] or 0))
            self.preco_venda.setValue(float(produto[5] or 0))
            self.estoque.setValue(float(produto[6] or 0))
            self.estoque_minimo.setValue(float(produto[7] or 0))

    def obter_dados(self):
        return {
            "codigo": self.codigo.text().strip(),
            "nome": self.nome.text().strip(),
            "categoria": self.categoria.text().strip(),
            "preco_compra": self.preco_compra.value(),
            "preco_venda": self.preco_venda.value(),
            "estoque": self.estoque.value(),
            "estoque_minimo": self.estoque_minimo.value(),
        }


# ============================================================
# CLIENTE
# ============================================================

class ClienteDialog(QDialog):

    def __init__(self, parent=None, cliente=None):
        super().__init__(parent)

        self.setWindowTitle(
            "Cadastrar Cliente"
            if cliente is None
            else "Editar Cliente"
        )

        layout = QFormLayout(self)

        self.nome = QLineEdit()
        self.cpf_cnpj = QLineEdit()
        self.telefone = QLineEdit()
        self.email = QLineEdit()

        layout.addRow("Nome:", self.nome)
        layout.addRow("CPF/CNPJ:", self.cpf_cnpj)
        layout.addRow("Telefone:", self.telefone)
        layout.addRow("E-mail:", self.email)

        botoes = QDialogButtonBox(
            QDialogButtonBox.Save |
            QDialogButtonBox.Cancel
        )

        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)

        layout.addRow(botoes)

        if cliente:
            self.nome.setText(cliente[1] or "")
            self.cpf_cnpj.setText(cliente[2] or "")
            self.telefone.setText(cliente[3] or "")
            self.email.setText(cliente[4] or "")

    def obter_dados(self):
        return {
            "nome": self.nome.text().strip(),
            "cpf_cnpj": self.cpf_cnpj.text().strip(),
            "telefone": self.telefone.text().strip(),
            "email": self.email.text().strip(),
        }


# ============================================================
# FORNECEDOR
# ============================================================

class FornecedorDialog(QDialog):

    def __init__(self, parent=None, fornecedor=None):
        super().__init__(parent)

        self.setWindowTitle(
            "Cadastrar Fornecedor"
            if fornecedor is None
            else "Editar Fornecedor"
        )

        layout = QFormLayout(self)

        self.nome = QLineEdit()
        self.cpf_cnpj = QLineEdit()
        self.telefone = QLineEdit()
        self.email = QLineEdit()
        self.endereco = QLineEdit()

        layout.addRow("Nome / Razão Social:", self.nome)
        layout.addRow("CPF/CNPJ:", self.cpf_cnpj)
        layout.addRow("Telefone:", self.telefone)
        layout.addRow("E-mail:", self.email)
        layout.addRow("Endereço:", self.endereco)

        botoes = QDialogButtonBox(
            QDialogButtonBox.Save |
            QDialogButtonBox.Cancel
        )

        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)

        layout.addRow(botoes)

        if fornecedor:
            self.nome.setText(fornecedor[1] or "")
            self.cpf_cnpj.setText(fornecedor[2] or "")
            self.telefone.setText(fornecedor[3] or "")
            self.email.setText(fornecedor[4] or "")
            self.endereco.setText(fornecedor[5] or "")

    def obter_dados(self):
        return {
            "nome": self.nome.text().strip(),
            "cpf_cnpj": self.cpf_cnpj.text().strip(),
            "telefone": self.telefone.text().strip(),
            "email": self.email.text().strip(),
            "endereco": self.endereco.text().strip(),
        }


# ============================================================
# MOVIMENTAÇÃO DE ESTOQUE
# ============================================================

class MovimentacaoDialog(QDialog):

    def __init__(self, database, tipo, parent=None):
        super().__init__(parent)

        self.database = database
        self.tipo = tipo

        self.setWindowTitle(
            "Entrada de Mercadoria"
            if tipo == "ENTRADA"
            else "Saída de Mercadoria"
        )

        layout = QFormLayout(self)

        self.produto = QComboBox()

        for produto in self.database.listar_produtos():
            self.produto.addItem(
                f"{produto[2]} | Estoque: {produto[6]}",
                produto[0]
            )

        self.quantidade = QDoubleSpinBox()
        self.quantidade.setMinimum(0.01)
        self.quantidade.setMaximum(999999999)
        self.quantidade.setDecimals(2)

        self.observacao = QLineEdit()

        layout.addRow("Produto:", self.produto)
        layout.addRow("Quantidade:", self.quantidade)
        layout.addRow("Observação:", self.observacao)

        botoes = QDialogButtonBox(
            QDialogButtonBox.Save |
            QDialogButtonBox.Cancel
        )

        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)

        layout.addRow(botoes)

    def obter_dados(self):
        return {
            "produto_id": self.produto.currentData(),
            "quantidade": self.quantidade.value(),
            "observacao": self.observacao.text().strip(),
        }


# ============================================================
# FLUXO DE CAIXA
# ============================================================

class MovimentoCaixaDialog(QDialog):

    def __init__(self, tipo, parent=None):
        super().__init__(parent)

        self.tipo = tipo

        self.setWindowTitle(
            "Nova Entrada"
            if tipo == "ENTRADA"
            else "Nova Saída"
        )

        layout = QFormLayout(self)

        self.descricao = QLineEdit()
        self.categoria = QLineEdit()

        self.valor = QDoubleSpinBox()
        self.valor.setMinimum(0.01)
        self.valor.setMaximum(999999999)
        self.valor.setDecimals(2)
        self.valor.setPrefix("R$ ")

        self.forma_pagamento = QComboBox()

        self.forma_pagamento.addItems([
            "Dinheiro",
            "PIX",
            "Cartão de Débito",
            "Cartão de Crédito",
            "Transferência",
            "Boleto",
            "Outro",
        ])

        self.observacao = QLineEdit()

        layout.addRow("Descrição:", self.descricao)
        layout.addRow("Categoria:", self.categoria)
        layout.addRow("Valor:", self.valor)
        layout.addRow(
            "Pagamento:",
            self.forma_pagamento
        )
        layout.addRow("Observação:", self.observacao)

        botoes = QDialogButtonBox(
            QDialogButtonBox.Save |
            QDialogButtonBox.Cancel
        )

        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)

        layout.addRow(botoes)

    def obter_dados(self):
        return {
            "tipo": self.tipo,
            "descricao": self.descricao.text().strip(),
            "categoria": self.categoria.text().strip(),
            "valor": self.valor.value(),
            "forma_pagamento":
                self.forma_pagamento.currentText(),
            "observacao":
                self.observacao.text().strip(),
        }


# ============================================================
# LOGIN
# ============================================================

class LoginWindow(QWidget):

    def __init__(self, database, callback):
        super().__init__()

        self.database = database
        self.callback = callback

        self.setWindowTitle("VS ERP - Login")
        self.resize(450, 500)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(60, 60, 60, 60)

        titulo = QLabel("VS ERP")
        titulo.setAlignment(Qt.AlignCenter)
        titulo.setStyleSheet(
            "font-size: 38px;"
            "font-weight: bold;"
            "color: #2563eb;"
        )

        subtitulo = QLabel(
            "Sistema de Gestão Comercial"
        )

        subtitulo.setAlignment(Qt.AlignCenter)

        self.usuario = QLineEdit()
        self.usuario.setPlaceholderText("Usuário")

        self.senha = QLineEdit()
        self.senha.setPlaceholderText("Senha")
        self.senha.setEchoMode(QLineEdit.Password)

        entrar = QPushButton("ENTRAR")
        entrar.clicked.connect(self.fazer_login)

        self.senha.returnPressed.connect(
            self.fazer_login
        )

        info = QLabel(
            "Usuário: admin\nSenha: 1234"
        )

        info.setAlignment(Qt.AlignCenter)

        layout.addStretch()
        layout.addWidget(titulo)
        layout.addWidget(subtitulo)
        layout.addSpacing(30)
        layout.addWidget(self.usuario)
        layout.addWidget(self.senha)
        layout.addWidget(entrar)
        layout.addWidget(info)
        layout.addStretch()

    def fazer_login(self):
        dados = self.database.verificar_login(
            self.usuario.text().strip(),
            self.senha.text()
        )

        if dados:
            self.callback(dados)
        else:
            QMessageBox.warning(
                self,
                "VS ERP",
                "Usuário ou senha inválidos."
            )


# ============================================================
# DASHBOARD
# ============================================================

class Dashboard(QWidget):

    def __init__(self, database):
        super().__init__()

        self.database = database

        layout = QVBoxLayout(self)

        titulo = QLabel("Dashboard")
        titulo.setStyleSheet(
            "font-size: 30px; font-weight: bold;"
        )

        layout.addWidget(titulo)

        self.cards = QHBoxLayout()

        self.vendas = self.criar_card(
            "Vendas Hoje"
        )

        self.produtos = self.criar_card(
            "Produtos"
        )

        self.clientes = self.criar_card(
            "Clientes"
        )

        self.saldo = self.criar_card(
            "Saldo Caixa"
        )

        layout.addLayout(self.cards)

        self.figure = Figure(
            figsize=(8, 4)
        )

        self.canvas = FigureCanvasQTAgg(
            self.figure
        )

        layout.addWidget(self.canvas)

        self.atualizar()

    def criar_card(self, titulo):
        frame = QFrame()

        frame.setStyleSheet(
            """
            QFrame {
                background-color: white;
                border-radius: 10px;
            }
            """
        )

        card = QVBoxLayout(frame)

        card.addWidget(
            QLabel(titulo)
        )

        valor = QLabel("0")

        valor.setStyleSheet(
            "font-size: 23px;"
            "font-weight: bold;"
        )

        card.addWidget(valor)

        self.cards.addWidget(frame)

        return valor

    def atualizar(self):
        self.vendas.setText(
            f"R$ "
            f"{self.database.total_vendas_hoje():.2f}"
        )

        self.produtos.setText(
            str(
                self.database.contar_produtos()
            )
        )

        self.clientes.setText(
            str(
                self.database.contar_clientes()
            )
        )

        self.saldo.setText(
            f"R$ "
            f"{self.database.saldo_caixa():.2f}"
        )

        self.atualizar_grafico()

    def atualizar_grafico(self):
        dados = (
            self.database
            .relatorio_vendas_por_dia(7)
        )

        dias = [
            item[0][:5]
            for item in dados
        ]

        valores = [
            item[1]
            for item in dados
        ]

        self.figure.clear()

        eixo = self.figure.add_subplot(111)

        eixo.plot(
            dias,
            valores,
            marker="o"
        )

        eixo.set_title(
            "Vendas dos últimos 7 dias"
        )

        eixo.set_ylabel("R$")

        eixo.tick_params(
            axis="x",
            rotation=30
        )

        self.figure.tight_layout()

        self.canvas.draw()


# ============================================================
# PRODUTOS
# ============================================================

class ProdutosPage(QWidget):

    def __init__(self, database, atualizar):
        super().__init__()

        self.database = database
        self.atualizar_callback = atualizar

        layout = QVBoxLayout(self)

        titulo = QLabel("Produtos")

        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        topo = QHBoxLayout()

        self.pesquisa = QLineEdit()

        self.pesquisa.setPlaceholderText(
            "Pesquisar produtos..."
        )

        self.pesquisa.textChanged.connect(
            self.carregar
        )

        novo = QPushButton("Novo Produto")

        novo.clicked.connect(
            self.novo
        )

        topo.addWidget(self.pesquisa)
        topo.addWidget(novo)

        layout.addLayout(topo)

        self.tabela = QTableWidget()
        self.tabela.setColumnCount(8)

        self.tabela.setHorizontalHeaderLabels([
            "ID",
            "Código",
            "Nome",
            "Categoria",
            "Preço Compra",
            "Preço Venda",
            "Estoque",
            "Mínimo",
        ])

        self.tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        self.tabela.setSelectionBehavior(
            QTableWidget.SelectRows
        )

        layout.addWidget(self.tabela)

        botoes = QHBoxLayout()

        editar = QPushButton("Editar")
        excluir = QPushButton("Excluir")

        editar.clicked.connect(self.editar)
        excluir.clicked.connect(self.excluir)

        botoes.addStretch()
        botoes.addWidget(editar)
        botoes.addWidget(excluir)

        layout.addLayout(botoes)

        self.carregar()

    def carregar(self):
        preencher_tabela(
            self.tabela,
            self.database.listar_produtos(
                self.pesquisa.text().strip()
            )
        )

    def selecionado(self):
        linha = self.tabela.currentRow()

        if linha < 0:
            QMessageBox.warning(
                self,
                "VS ERP",
                "Selecione um produto."
            )
            return None

        return int(
            self.tabela.item(
                linha,
                0
            ).text()
        )

    def novo(self):
        dialog = ProdutoDialog(self)

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        if not dados["nome"]:
            return

        try:
            self.database.cadastrar_produto(
                dados["codigo"],
                dados["nome"],
                dados["categoria"],
                dados["preco_compra"],
                dados["preco_venda"],
                dados["estoque"],
                dados["estoque_minimo"]
            )

            self.carregar()
            self.atualizar_callback()

        except sqlite3.IntegrityError:
            QMessageBox.warning(
                self,
                "VS ERP",
                "Código já cadastrado."
            )

    def editar(self):
        produto_id = self.selecionado()

        if produto_id is None:
            return

        produto = self.database.buscar_produto_por_id(
            produto_id
        )

        dialog = ProdutoDialog(
            self,
            produto
        )

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        self.database.atualizar_produto(
            produto_id,
            dados["codigo"],
            dados["nome"],
            dados["categoria"],
            dados["preco_compra"],
            dados["preco_venda"],
            dados["estoque"],
            dados["estoque_minimo"]
        )

        self.carregar()
        self.atualizar_callback()

    def excluir(self):
        produto_id = self.selecionado()

        if produto_id is None:
            return

        try:
            self.database.excluir_produto(
                produto_id
            )

            self.carregar()
            self.atualizar_callback()

        except sqlite3.IntegrityError:
            QMessageBox.warning(
                self,
                "VS ERP",
                "O produto possui histórico "
                "e não pode ser excluído."
            )


# ============================================================
# CLIENTES
# ============================================================

class ClientesPage(QWidget):

    def __init__(self, database, atualizar):
        super().__init__()

        self.database = database
        self.atualizar_callback = atualizar

        layout = QVBoxLayout(self)

        titulo = QLabel("Clientes")
        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        topo = QHBoxLayout()

        self.pesquisa = QLineEdit()

        self.pesquisa.setPlaceholderText(
            "Pesquisar clientes..."
        )

        self.pesquisa.textChanged.connect(
            self.carregar
        )

        novo = QPushButton("Novo Cliente")

        novo.clicked.connect(
            self.novo
        )

        topo.addWidget(self.pesquisa)
        topo.addWidget(novo)

        layout.addLayout(topo)

        self.tabela = QTableWidget()
        self.tabela.setColumnCount(5)

        self.tabela.setHorizontalHeaderLabels([
            "ID",
            "Nome",
            "CPF/CNPJ",
            "Telefone",
            "E-mail",
        ])

        self.tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        self.tabela.setSelectionBehavior(
            QTableWidget.SelectRows
        )

        layout.addWidget(self.tabela)

        botoes = QHBoxLayout()

        editar = QPushButton("Editar")
        excluir = QPushButton("Excluir")

        editar.clicked.connect(self.editar)
        excluir.clicked.connect(self.excluir)

        botoes.addStretch()
        botoes.addWidget(editar)
        botoes.addWidget(excluir)

        layout.addLayout(botoes)

        self.carregar()

    def carregar(self):
        preencher_tabela(
            self.tabela,
            self.database.listar_clientes(
                self.pesquisa.text().strip()
            )
        )

    def selecionado(self):
        linha = self.tabela.currentRow()

        if linha < 0:
            QMessageBox.warning(
                self,
                "VS ERP",
                "Selecione um cliente."
            )
            return None

        return int(
            self.tabela.item(
                linha,
                0
            ).text()
        )

    def novo(self):
        dialog = ClienteDialog(self)

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        if not dados["nome"]:
            return

        self.database.cadastrar_cliente(
            dados["nome"],
            dados["cpf_cnpj"],
            dados["telefone"],
            dados["email"]
        )

        self.carregar()
        self.atualizar_callback()

    def editar(self):
        cliente_id = self.selecionado()

        if cliente_id is None:
            return

        cliente = self.database.buscar_cliente_por_id(
            cliente_id
        )

        dialog = ClienteDialog(
            self,
            cliente
        )

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        self.database.atualizar_cliente(
            cliente_id,
            dados["nome"],
            dados["cpf_cnpj"],
            dados["telefone"],
            dados["email"]
        )

        self.carregar()

    def excluir(self):
        cliente_id = self.selecionado()

        if cliente_id is None:
            return

        self.database.excluir_cliente(
            cliente_id
        )

        self.carregar()
        self.atualizar_callback()


# ============================================================
# FORNECEDORES
# ============================================================

class FornecedoresPage(QWidget):

    def __init__(self, database, atualizar):
        super().__init__()

        self.database = database
        self.atualizar_callback = atualizar

        layout = QVBoxLayout(self)

        titulo = QLabel("Fornecedores")
        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        topo = QHBoxLayout()

        self.pesquisa = QLineEdit()

        self.pesquisa.setPlaceholderText(
            "Pesquisar fornecedores..."
        )

        self.pesquisa.textChanged.connect(
            self.carregar
        )

        novo = QPushButton(
            "Novo Fornecedor"
        )

        novo.clicked.connect(
            self.novo
        )

        topo.addWidget(self.pesquisa)
        topo.addWidget(novo)

        layout.addLayout(topo)

        self.tabela = QTableWidget()
        self.tabela.setColumnCount(6)

        self.tabela.setHorizontalHeaderLabels([
            "ID",
            "Nome",
            "CPF/CNPJ",
            "Telefone",
            "E-mail",
            "Endereço",
        ])

        self.tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        self.tabela.setSelectionBehavior(
            QTableWidget.SelectRows
        )

        layout.addWidget(self.tabela)

        botoes = QHBoxLayout()

        editar = QPushButton("Editar")
        excluir = QPushButton("Excluir")

        editar.clicked.connect(self.editar)
        excluir.clicked.connect(self.excluir)

        botoes.addStretch()
        botoes.addWidget(editar)
        botoes.addWidget(excluir)

        layout.addLayout(botoes)

        self.carregar()

    def carregar(self):
        preencher_tabela(
            self.tabela,
            self.database.listar_fornecedores(
                self.pesquisa.text().strip()
            )
        )

    def selecionado(self):
        linha = self.tabela.currentRow()

        if linha < 0:
            return None

        return int(
            self.tabela.item(
                linha,
                0
            ).text()
        )

    def novo(self):
        dialog = FornecedorDialog(self)

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        if not dados["nome"]:
            return

        self.database.cadastrar_fornecedor(
            dados["nome"],
            dados["cpf_cnpj"],
            dados["telefone"],
            dados["email"],
            dados["endereco"]
        )

        self.carregar()
        self.atualizar_callback()

    def editar(self):
        fornecedor_id = self.selecionado()

        if fornecedor_id is None:
            return

        fornecedor = (
            self.database
            .buscar_fornecedor_por_id(
                fornecedor_id
            )
        )

        dialog = FornecedorDialog(
            self,
            fornecedor
        )

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        self.database.atualizar_fornecedor(
            fornecedor_id,
            dados["nome"],
            dados["cpf_cnpj"],
            dados["telefone"],
            dados["email"],
            dados["endereco"]
        )

        self.carregar()

    def excluir(self):
        fornecedor_id = self.selecionado()

        if fornecedor_id is None:
            return

        self.database.excluir_fornecedor(
            fornecedor_id
        )

        self.carregar()
        self.atualizar_callback()


# ============================================================
# ESTOQUE
# ============================================================

class EstoquePage(QWidget):

    def __init__(self, database):
        super().__init__()

        self.database = database

        layout = QVBoxLayout(self)

        titulo = QLabel(
            "Controle de Estoque"
        )

        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        self.tabela = QTableWidget()
        self.tabela.setColumnCount(6)

        self.tabela.setHorizontalHeaderLabels([
            "Código",
            "Produto",
            "Categoria",
            "Estoque Atual",
            "Estoque Mínimo",
            "Situação",
        ])

        self.tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        layout.addWidget(self.tabela)

        self.carregar()

    def carregar(self):
        produtos = self.database.listar_produtos()

        dados = []

        for produto in produtos:
            estoque = float(
                produto[6] or 0
            )

            minimo = float(
                produto[7] or 0
            )

            situacao = (
                "ESTOQUE BAIXO"
                if estoque <= minimo
                else "NORMAL"
            )

            dados.append((
                produto[1],
                produto[2],
                produto[3],
                estoque,
                minimo,
                situacao
            ))

        preencher_tabela(
            self.tabela,
            dados
        )


# ============================================================
# CAIXA / PDV
# ============================================================

class CaixaPage(QWidget):

    def __init__(self, database, atualizar):
        super().__init__()

        self.database = database
        self.atualizar_callback = atualizar
        self.itens = []

        layout = QVBoxLayout(self)

        titulo = QLabel("Caixa / PDV")

        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        topo = QHBoxLayout()

        self.produto = QComboBox()

        self.quantidade = QDoubleSpinBox()
        self.quantidade.setMinimum(0.01)
        self.quantidade.setMaximum(999999)
        self.quantidade.setValue(1)

        adicionar = QPushButton(
            "Adicionar"
        )

        adicionar.clicked.connect(
            self.adicionar
        )

        topo.addWidget(self.produto)
        topo.addWidget(self.quantidade)
        topo.addWidget(adicionar)

        layout.addLayout(topo)

        self.tabela = QTableWidget()
        self.tabela.setColumnCount(4)

        self.tabela.setHorizontalHeaderLabels([
            "Produto",
            "Quantidade",
            "Preço",
            "Subtotal",
        ])

        self.tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        layout.addWidget(self.tabela)

        self.total = QLabel(
            "TOTAL: R$ 0,00"
        )

        self.total.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(self.total)

        self.pagamento = QComboBox()

        self.pagamento.addItems([
            "Dinheiro",
            "PIX",
            "Cartão de Débito",
            "Cartão de Crédito",
        ])

        finalizar = QPushButton(
            "FINALIZAR VENDA"
        )

        finalizar.clicked.connect(
            self.finalizar
        )

        layout.addWidget(self.pagamento)
        layout.addWidget(finalizar)

        self.carregar_produtos()

    def carregar_produtos(self):
        self.produto.clear()

        for produto in self.database.listar_produtos():
            self.produto.addItem(
                f"{produto[2]} | "
                f"Estoque: {produto[6]}",
                produto[0]
            )

    def adicionar(self):
        produto_id = self.produto.currentData()

        if produto_id is None:
            return

        produto = (
            self.database
            .buscar_produto_por_id(
                produto_id
            )
        )

        quantidade = (
            self.quantidade.value()
        )

        estoque = float(
            produto[6] or 0
        )

        total_carrinho = sum(
            item["quantidade"]
            for item in self.itens
            if item["produto_id"]
            == produto_id
        )

        if (
            quantidade
            +
            total_carrinho
            >
            estoque
        ):
            QMessageBox.warning(
                self,
                "VS ERP",
                "Estoque insuficiente."
            )
            return

        self.itens.append({
            "produto_id": produto_id,
            "nome": produto[2],
            "quantidade": quantidade,
            "preco": float(
                produto[5] or 0
            ),
        })

        self.atualizar_carrinho()

    def atualizar_carrinho(self):
        dados = []

        total = 0

        for item in self.itens:
            subtotal = (
                item["quantidade"]
                *
                item["preco"]
            )

            total += subtotal

            dados.append((
                item["nome"],
                item["quantidade"],
                f"R$ {item['preco']:.2f}",
                f"R$ {subtotal:.2f}"
            ))

        preencher_tabela(
            self.tabela,
            dados
        )

        self.total.setText(
            f"TOTAL: R$ {total:.2f}"
        )

    def finalizar(self):
        if not self.itens:
            QMessageBox.warning(
                self,
                "VS ERP",
                "Adicione produtos à venda."
            )
            return

        itens = [
            {
                "produto_id":
                    item["produto_id"],

                "quantidade":
                    item["quantidade"],
            }
            for item in self.itens
        ]

        try:
            venda_id, total = (
                self.database
                .finalizar_venda(
                    itens,
                    self.pagamento.currentText()
                )
            )

            QMessageBox.information(
                self,
                "VS ERP",
                f"Venda #{venda_id} "
                f"finalizada.\n"
                f"Total: R$ {total:.2f}"
            )

            self.itens.clear()

            self.atualizar_carrinho()

            self.atualizar_callback()

        except ValueError as erro:
            QMessageBox.warning(
                self,
                "VS ERP",
                str(erro)
            )


# ============================================================
# VENDAS
# ============================================================

class VendasPage(QWidget):

    def __init__(self, database):
        super().__init__()

        self.database = database

        layout = QVBoxLayout(self)

        titulo = QLabel("Vendas")

        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        self.tabela = QTableWidget()
        self.tabela.setColumnCount(5)

        self.tabela.setHorizontalHeaderLabels([
            "Venda",
            "Data/Hora",
            "Total",
            "Pagamento",
            "Observação",
        ])

        self.tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        layout.addWidget(self.tabela)

        self.carregar()

    def carregar(self):
        vendas = (
            self.database
            .listar_vendas()
        )

        dados = []

        for venda in vendas:
            dados.append((
                venda[0],
                venda[1],
                f"R$ {venda[2]:.2f}",
                venda[3],
                venda[4]
            ))

        preencher_tabela(
            self.tabela,
            dados
        )


# ============================================================
# FLUXO DE CAIXA
# ============================================================

class FluxoCaixaPage(QWidget):

    def __init__(self, database, atualizar):
        super().__init__()

        self.database = database
        self.atualizar_callback = atualizar

        layout = QVBoxLayout(self)

        titulo = QLabel(
            "Fluxo de Caixa"
        )

        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        resumo = QHBoxLayout()

        self.entradas = QLabel()
        self.saidas = QLabel()
        self.saldo = QLabel()

        for label in (
            self.entradas,
            self.saidas,
            self.saldo
        ):
            label.setStyleSheet(
                "background: white;"
                "padding: 15px;"
                "font-size: 18px;"
                "font-weight: bold;"
            )

            resumo.addWidget(label)

        layout.addLayout(resumo)

        botoes = QHBoxLayout()

        entrada = QPushButton(
            "Nova Entrada"
        )

        saida = QPushButton(
            "Nova Saída"
        )

        entrada.clicked.connect(
            lambda:
            self.novo_movimento(
                "ENTRADA"
            )
        )

        saida.clicked.connect(
            lambda:
            self.novo_movimento(
                "SAÍDA"
            )
        )

        botoes.addWidget(entrada)
        botoes.addWidget(saida)
        botoes.addStretch()

        layout.addLayout(botoes)

        self.tabela = QTableWidget()
        self.tabela.setColumnCount(8)

        self.tabela.setHorizontalHeaderLabels([
            "ID",
            "Tipo",
            "Descrição",
            "Categoria",
            "Valor",
            "Pagamento",
            "Data/Hora",
            "Observação",
        ])

        self.tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        layout.addWidget(self.tabela)

        self.carregar()

    def novo_movimento(self, tipo):
        dialog = MovimentoCaixaDialog(
            tipo,
            self
        )

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        try:
            self.database.cadastrar_movimento_caixa(
                dados["tipo"],
                dados["descricao"],
                dados["categoria"],
                dados["valor"],
                dados["forma_pagamento"],
                dados["observacao"]
            )

            self.carregar()
            self.atualizar_callback()

        except ValueError as erro:
            QMessageBox.warning(
                self,
                "VS ERP",
                str(erro)
            )

    def carregar(self):
        movimentos = (
            self.database
            .listar_fluxo_caixa()
        )

        dados = []

        for movimento in movimentos:
            dados.append((
                movimento[0],
                movimento[1],
                movimento[2],
                movimento[3],
                f"R$ {movimento[4]:.2f}",
                movimento[5],
                movimento[6],
                movimento[7]
            ))

        preencher_tabela(
            self.tabela,
            dados
        )

        self.entradas.setText(
            "Entradas: "
            f"R$ "
            f"{self.database.total_entradas_caixa():.2f}"
        )

        self.saidas.setText(
            "Saídas: "
            f"R$ "
            f"{self.database.total_saidas_caixa():.2f}"
        )

        self.saldo.setText(
            "Saldo: "
            f"R$ "
            f"{self.database.saldo_caixa():.2f}"
        )


# ============================================================
# RELATÓRIOS + EXCEL + GRÁFICOS
# ============================================================

class RelatoriosPage(QWidget):

    def __init__(self, database):
        super().__init__()

        self.database = database

        layout = QVBoxLayout(self)

        titulo = QLabel(
            "Relatórios e Exportações"
        )

        titulo.setStyleSheet(
            "font-size: 28px;"
            "font-weight: bold;"
        )

        layout.addWidget(titulo)

        botoes = QHBoxLayout()

        exportar = QPushButton(
            "Exportar Tudo para Excel"
        )

        exportar.clicked.connect(
            self.exportar_excel
        )

        atualizar = QPushButton(
            "Atualizar Relatórios"
        )

        atualizar.clicked.connect(
            self.atualizar
        )

        botoes.addWidget(exportar)
        botoes.addWidget(atualizar)
        botoes.addStretch()

        layout.addLayout(botoes)

        self.abas = QTabWidget()

        self.tabela_vendas = QTableWidget()

        self.tabela_estoque = QTableWidget()

        self.tabela_produtos = QTableWidget()

        self.grafico_widget = QWidget()

        self.criar_abas()

        layout.addWidget(self.abas)

        self.atualizar()

    def configurar_tabela(
        self,
        tabela,
        cabecalhos
    ):
        tabela.setColumnCount(
            len(cabecalhos)
        )

        tabela.setHorizontalHeaderLabels(
            cabecalhos
        )

        tabela.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

    def criar_abas(self):
        self.configurar_tabela(
            self.tabela_vendas,
            [
                "Venda",
                "Data/Hora",
                "Total",
                "Pagamento",
                "Observação",
            ]
        )

        self.configurar_tabela(
            self.tabela_estoque,
            [
                "Código",
                "Produto",
                "Categoria",
                "Preço Compra",
                "Preço Venda",
                "Estoque",
                "Mínimo",
            ]
        )

        self.configurar_tabela(
            self.tabela_produtos,
            [
                "Código",
                "Produto",
                "Qtd. Vendida",
                "Valor Vendido",
            ]
        )

        self.abas.addTab(
            self.tabela_vendas,
            "Vendas"
        )

        self.abas.addTab(
            self.tabela_estoque,
            "Estoque"
        )

        self.abas.addTab(
            self.tabela_produtos,
            "Mais Vendidos"
        )

        grafico_layout = QVBoxLayout(
            self.grafico_widget
        )

        self.figure = Figure(
            figsize=(9, 5)
        )

        self.canvas = FigureCanvasQTAgg(
            self.figure
        )

        grafico_layout.addWidget(
            self.canvas
        )

        self.abas.addTab(
            self.grafico_widget,
            "Gráfico de Vendas"
        )

    def atualizar(self):
        vendas = (
            self.database
            .relatorio_vendas_completo()
        )

        vendas_formatadas = []

        for venda in vendas:
            vendas_formatadas.append((
                venda[0],
                venda[1],
                f"R$ {venda[2]:.2f}",
                venda[3],
                venda[4]
            ))

        preencher_tabela(
            self.tabela_vendas,
            vendas_formatadas
        )

        preencher_tabela(
            self.tabela_estoque,
            self.database
            .relatorio_estoque_atual()
        )

        mais_vendidos = (
            self.database
            .relatorio_produtos_mais_vendidos(
                10
            )
        )

        dados_produtos = []

        for item in mais_vendidos:
            dados_produtos.append((
                item[0],
                item[1],
                item[2],
                f"R$ {item[3]:.2f}"
            ))

        preencher_tabela(
            self.tabela_produtos,
            dados_produtos
        )

        self.atualizar_grafico()

    def atualizar_grafico(self):
        dados = (
            self.database
            .relatorio_vendas_por_dia(30)
        )

        datas = [
            item[0][:5]
            for item in dados
        ]

        valores = [
            item[1]
            for item in dados
        ]

        self.figure.clear()

        eixo = self.figure.add_subplot(111)

        eixo.plot(
            datas,
            valores,
            marker="o"
        )

        eixo.set_title(
            "Vendas dos últimos 30 dias"
        )

        eixo.set_xlabel("Data")

        eixo.set_ylabel("Valor vendido (R$)")

        eixo.tick_params(
            axis="x",
            rotation=60
        )

        self.figure.tight_layout()

        self.canvas.draw()

    def exportar_excel(self):
        nome_padrao = (
            "VS_ERP_Relatorio_"
            +
            datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            +
            ".xlsx"
        )

        caminho, _ = (
            QFileDialog
            .getSaveFileName(
                self,
                "Salvar relatório Excel",
                str(
                    EXPORT_DIR /
                    nome_padrao
                ),
                "Arquivo Excel (*.xlsx)"
            )
        )

        if not caminho:
            return

        try:
            workbook = Workbook()

            aba_dashboard = workbook.active

            aba_dashboard.title = (
                "Dashboard"
            )

            resumo = (
                self.database
                .relatorio_resumo_financeiro()
            )

            aba_dashboard.append([
                "VS ERP - Resumo Financeiro"
            ])

            aba_dashboard[
                "A1"
            ].font = Font(
                bold=True,
                size=16
            )

            aba_dashboard.append([])
            aba_dashboard.append([
                "Indicador",
                "Valor"
            ])

            aba_dashboard.append([
                "Total de Vendas",
                resumo["vendas_total"]
            ])

            aba_dashboard.append([
                "Entradas de Caixa",
                resumo["entradas"]
            ])

            aba_dashboard.append([
                "Saídas de Caixa",
                resumo["saidas"]
            ])

            aba_dashboard.append([
                "Saldo Atual",
                resumo["saldo"]
            ])

            aba_dashboard.append([
                "Vendas Hoje",
                resumo["vendas_hoje"]
            ])

            aba_produtos = (
                workbook.create_sheet(
                    "Produtos"
                )
            )

            aba_produtos.append([
                "Código",
                "Nome",
                "Categoria",
                "Preço Compra",
                "Preço Venda",
                "Estoque",
                "Estoque Mínimo",
            ])

            for produto in (
                self.database
                .relatorio_estoque_atual()
            ):
                aba_produtos.append(
                    list(produto)
                )

            aba_vendas = (
                workbook.create_sheet(
                    "Vendas"
                )
            )

            aba_vendas.append([
                "ID",
                "Data/Hora",
                "Total",
                "Pagamento",
                "Observação",
            ])

            for venda in (
                self.database
                .relatorio_vendas_completo()
            ):
                aba_vendas.append(
                    list(venda)
                )

            aba_fluxo = (
                workbook.create_sheet(
                    "Fluxo de Caixa"
                )
            )

            aba_fluxo.append([
                "ID",
                "Tipo",
                "Descrição",
                "Categoria",
                "Valor",
                "Pagamento",
                "Data/Hora",
                "Observação",
            ])

            for movimento in (
                self.database
                .relatorio_fluxo_caixa_completo()
            ):
                aba_fluxo.append(
                    list(movimento)
                )

            aba_clientes = (
                workbook.create_sheet(
                    "Clientes"
                )
            )

            aba_clientes.append([
                "ID",
                "Nome",
                "CPF/CNPJ",
                "Telefone",
                "E-mail",
            ])

            for cliente in (
                self.database
                .relatorio_clientes()
            ):
                aba_clientes.append(
                    list(cliente)
                )

            aba_fornecedores = (
                workbook.create_sheet(
                    "Fornecedores"
                )
            )

            aba_fornecedores.append([
                "ID",
                "Nome",
                "CPF/CNPJ",
                "Telefone",
                "E-mail",
                "Endereço",
            ])

            for fornecedor in (
                self.database
                .relatorio_fornecedores()
            ):
                aba_fornecedores.append(
                    list(fornecedor)
                )

            aba_grafico = (
                workbook.create_sheet(
                    "Vendas por Dia"
                )
            )

            aba_grafico.append([
                "Data",
                "Total"
            ])

            dados_vendas = (
                self.database
                .relatorio_vendas_por_dia(
                    30
                )
            )

            for dia, total in dados_vendas:
                aba_grafico.append([
                    dia,
                    total
                ])

            grafico = LineChart()

            grafico.title = (
                "Vendas dos Últimos 30 Dias"
            )

            grafico.y_axis.title = (
                "Valor Vendido"
            )

            grafico.x_axis.title = (
                "Data"
            )

            dados_ref = Reference(
                aba_grafico,
                min_col=2,
                min_row=1,
                max_row=(
                    len(dados_vendas)
                    +
                    1
                )
            )

            categorias_ref = Reference(
                aba_grafico,
                min_col=1,
                min_row=2,
                max_row=(
                    len(dados_vendas)
                    +
                    1
                )
            )

            grafico.add_data(
                dados_ref,
                titles_from_data=True
            )

            grafico.set_categories(
                categorias_ref
            )

            aba_grafico.add_chart(
                grafico,
                "D2"
            )

            for aba in workbook.worksheets:
                for celula in aba[1]:
                    celula.font = Font(
                        bold=True
                    )

                    celula.alignment = (
                        Alignment(
                            horizontal="center"
                        )
                    )

                for coluna in (
                    aba.columns
                ):
                    maior = 0

                    letra = (
                        coluna[0]
                        .column_letter
                    )

                    for celula in coluna:
                        valor = (
                            ""
                            if celula.value is None
                            else str(
                                celula.value
                            )
                        )

                        maior = max(
                            maior,
                            len(valor)
                        )

                    aba.column_dimensions[
                        letra
                    ].width = min(
                        maior + 3,
                        40
                    )

            workbook.save(
                caminho
            )

            QMessageBox.information(
                self,
                "VS ERP",
                "Relatório Excel gerado "
                "com sucesso.\n\n"
                f"{caminho}"
            )

        except Exception as erro:
            QMessageBox.critical(
                self,
                "VS ERP",
                "Não foi possível gerar "
                "o relatório.\n\n"
                f"{erro}"
            )


# ============================================================
# JANELA PRINCIPAL
# ============================================================

class MainWindow(QMainWindow):

    def __init__(
        self,
        database,
        usuario
    ):
        super().__init__()

        self.database = database
        self.usuario = usuario

        self.setWindowTitle(
            "VS ERP - Sistema de Gestão Comercial"
        )

        self.resize(
            1400,
            850
        )

        central = QWidget()

        self.setCentralWidget(
            central
        )

        layout = QHBoxLayout(
            central
        )

        layout.setContentsMargins(
            0,
            0,
            0,
            0
        )

        layout.setSpacing(0)

        menu = QFrame()

        menu.setFixedWidth(
            230
        )

        menu.setStyleSheet(
            """
            QFrame {
                background-color: #0f172a;
            }

            QLabel {
                color: white;
            }

            QPushButton {
                color: white;
                background-color: transparent;
                border: none;
                text-align: left;
                padding: 11px;
            }

            QPushButton:hover {
                background-color: #1e293b;
            }
            """
        )

        menu_layout = QVBoxLayout(
            menu
        )

        logo = QLabel("VS ERP")

        logo.setStyleSheet(
            "font-size: 25px;"
            "font-weight: bold;"
            "padding: 20px;"
        )

        menu_layout.addWidget(
            logo
        )

        self.paginas = (
            QStackedWidget()
        )

        self.dashboard = Dashboard(
            database
        )

        self.caixa = CaixaPage(
            database,
            self.atualizar_telas
        )

        self.produtos = ProdutosPage(
            database,
            self.atualizar_telas
        )

        self.estoque = EstoquePage(
            database
        )

        self.clientes = ClientesPage(
            database,
            self.atualizar_telas
        )

        self.fornecedores = (
            FornecedoresPage(
                database,
                self.atualizar_telas
            )
        )

        self.vendas = VendasPage(
            database
        )

        self.fluxo = FluxoCaixaPage(
            database,
            self.atualizar_telas
        )

        self.relatorios = RelatoriosPage(
        
            database
        )
        self.configuracoes_impressora = ConfiguracoesImpressoraPage()
        
        paginas = [

            self.dashboard,
            self.caixa,
            self.produtos,
            self.estoque,
            self.clientes,
            self.fornecedores,
            self.vendas,
            self.fluxo,
            self.relatorios,
            self.configuracoes_impressora,
        ]

        for pagina in paginas:
            self.paginas.addWidget(
                pagina
            )

        botoes = [
            (
                "Dashboard",
                self.dashboard
            ),
            (
                "Caixa / PDV",
                self.caixa
            ),
            (
                "Produtos",
                self.produtos
            ),
            (
                "Estoque",
                self.estoque
            ),
            (
                "Clientes",
                self.clientes
            ),
            (
                "Fornecedores",
                self.fornecedores
            ),
            (
                "Vendas",
                self.vendas
            ),
            (
                "Fluxo de Caixa",
                self.fluxo
            ),
            (
                "Relatórios",
                self.relatorios
            ),
        (
    "Configurações",
    self.configuracoes_impressora
),
]

        for texto, pagina in botoes:
            botao = QPushButton(
                texto
            )

            botao.clicked.connect(
                lambda checked=False,
                p=pagina:
                self.abrir(p)
            )

            menu_layout.addWidget(
                botao
            )

        entrada = QPushButton(
            "Entrada de Estoque"
        )

        saida = QPushButton(
            "Saída de Estoque"
        )

        entrada.clicked.connect(
            self.entrada_estoque
        )

        saida.clicked.connect(
            self.saida_estoque
        )

        menu_layout.addWidget(
            entrada
        )

        menu_layout.addWidget(
            saida
        )

        menu_layout.addStretch()

        usuario_label = QLabel(
            f"{self.usuario[2]}\n"
            f"{self.usuario[3]}"
        )

        usuario_label.setStyleSheet(
            "padding: 15px;"
            "color: #94a3b8;"
        )

        menu_layout.addWidget(
            usuario_label
        )

        layout.addWidget(menu)
        layout.addWidget(
            self.paginas
        )

    def abrir(self, pagina):
        self.atualizar_telas()

        self.paginas.setCurrentWidget(
            pagina
        )

    def atualizar_telas(self):
        self.dashboard.atualizar()
        self.produtos.carregar()
        self.estoque.carregar()
        self.clientes.carregar()
        self.fornecedores.carregar()
        self.vendas.carregar()
        self.fluxo.carregar()
        self.caixa.carregar_produtos()
        self.relatorios.atualizar()

    def entrada_estoque(self):
        if self.database.contar_produtos() == 0:
            QMessageBox.warning(
                self,
                "VS ERP",
                "Cadastre um produto primeiro."
            )
            return

        dialog = MovimentacaoDialog(
            self.database,
            "ENTRADA",
            self
        )

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        try:
            self.database.registrar_entrada(
                dados["produto_id"],
                dados["quantidade"],
                dados["observacao"]
            )

            self.atualizar_telas()

            QMessageBox.information(
                self,
                "VS ERP",
                "Entrada registrada."
            )

        except ValueError as erro:
            QMessageBox.warning(
                self,
                "VS ERP",
                str(erro)
            )

    def saida_estoque(self):
        if self.database.contar_produtos() == 0:
            QMessageBox.warning(
                self,
                "VS ERP",
                "Cadastre um produto primeiro."
            )
            return

        dialog = MovimentacaoDialog(
            self.database,
            "SAÍDA",
            self
        )

        if dialog.exec() != QDialog.Accepted:
            return

        dados = dialog.obter_dados()

        try:
            self.database.registrar_saida(
                dados["produto_id"],
                dados["quantidade"],
                dados["observacao"]
            )

            self.atualizar_telas()

            QMessageBox.information(
                self,
                "VS ERP",
                "Saída registrada."
            )

        except ValueError as erro:
            QMessageBox.warning(
                self,
                "VS ERP",
                str(erro)
            )


# ============================================================
# SISTEMA
# ============================================================

class Sistema:

    def __init__(self):
        self.app = QApplication(
            sys.argv
        )

        self.app.setStyleSheet(
            """
            QWidget {
                font-family: Segoe UI;
                background-color: #f1f5f9;
                color: #0f172a;
            }

            QLineEdit,
            QComboBox,
            QDoubleSpinBox {
                background-color: white;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                padding: 7px;
            }

            QPushButton {
                padding: 8px 14px;
            }

            QTableWidget {
                background-color: white;
                border: 1px solid #cbd5e1;
            }

            QHeaderView::section {
                background-color: #e2e8f0;
                padding: 8px;
                font-weight: bold;
            }

            QTabWidget::pane {
                border: 1px solid #cbd5e1;
                background: white;
            }
            """
        )

        self.database = Database()

        self.login = LoginWindow(
            self.database,
            self.abrir_sistema
        )

        self.login.show()

    def abrir_sistema(
        self,
        usuario
    ):
        self.janela = MainWindow(
            self.database,
            usuario
        )

        self.janela.show()

        self.login.close()

    def executar(self):
        sys.exit(
            self.app.exec()
        )


if __name__ == "__main__":
    sistema = Sistema()
    sistema.executar()